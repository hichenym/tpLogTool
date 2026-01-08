import os
import sys
import time
import datetime
import json
import base64
import hashlib
import random
import ddddocr
import requests
import threading
from openpyxl import load_workbook

class Console_Admin_System:
    def __init__(self, env, username, password):
        """
        初始化相关参数
        生产环境使用wake与wake_many_times进行,生产环境需要鉴权，不能直接访问最底层的接口
        测试环境可以使用ake与wake_many_times进行，也可以使用wake_test与wake_many_times_test进行，测试环境公司内网不需要鉴权，直接可以调用最底层的接口
        :param env: pro：生产环境/test:测试环境
        :param username: 用户名
        :param password：密码
        """
        self.username = username
        self.password = password
        if env == 'pro':
            self.host = 'console.seetong.com'
        else:
            self.host = 'console-test.seetong.com'
        self.token, self.refresh_token = self.token_get()

    def refresh_token_wrap():
        """
        token过期后重新获取新token
        """

        def _refresh_token_wrap(func):
            def __refresh_token_wrap(self, *args, **kargs):

                info = func(self, *args, **kargs)
                if info[1] == 401:
                    print('token过期，重新获取token')
                    self.token, self.refresh_token = self.refresh_token_get(self.refresh_token)
                    info = func(self, *args, **kargs)
                return info

            return __refresh_token_wrap

        return _refresh_token_wrap

    def login_retry(times=3, interval=1):
        """
        captcha函数识别验证码可能不准，如获取token失败，可以尝试重新获取验证码
        :param app: 重试次数，默认重试3次
        :param app: 重试间隔，默认重试间隔1s
        """

        def _login_retry(func):
            def __login_retry(*args, **kargs):
                for index in range(times):
                    res = func(*args, **kargs)
                    if res[0]:
                        print(f'第{index+1}次调用{func.__name__}验证成功')
                        break
                    else:
                        print(f'第{index+1}次调用{func.__name__}验证失败')
                        if index < times - 1:
                            time.sleep(interval)
                if res:
                    return res
                else:
                    print(f'调用{func.__name__}验证失败，退出执行')
                    sys.exit()

            return __login_retry

        return _login_retry

    def captcha(self):
        """
        获取图片验证码captcha_code以及对应captcha_key
        """
        api_path = '/api/seetong-auth/oauth/captcha'
        url = f'https://{self.host}{api_path}'
        r = requests.get(url, verify=False)
        res = r.content.decode()
        res_dict = json.loads(res)

        image = res_dict.get('image')
        captcha_key = res_dict.get('key')

        # image中data:image/png;base64,后面的是图片的二进制base64编码数据
        img_data_base64 = image.split('data:image/png;base64,')[-1]
        img_data = base64.b64decode(img_data_base64.encode())

        # 根据图片数据识别图片二维码
        ocr = ddddocr.DdddOcr(show_ad=False)
        captcha_code = ocr.classification(img_data)
        return captcha_key, captcha_code

    @login_retry(3, 1)
    def token_get(self):
        """
        获取登录token用于调用其他函数鉴权
        :param app: 重试次数，默认重试3次
        :param app: 重试间隔，默认重试间隔1s
        """
        captcha_key, captcha_code = self.captcha()
        api_path = '/api/seetong-auth/oauth/token'
        url = f'https://{self.host}{api_path}'
        headers = {
            "User-Agent": "topsee/product-retrospect",
            "Content-Type": "application/json",
            "Captcha-Code": captcha_code,
            "Captcha-Key": captcha_key,
            "Authorization": "Basic c2VldG9uZ19jbG91ZF9hZG1pbjpzZWV0b25nX2Nsb3VkX2FkbWluX3NlY3JldA==",
            "Tenant-Id": "000000",
        }
        params = {
            "tenantId": "000000",
            "username": self.username,
            "password": hashlib.md5(self.password.encode()).hexdigest(),
            "grant_type": "captcha",
            "scope": "all",
            "type": "account",
        }
        r = requests.post(url, params=params, headers=headers, verify=False)
        res = r.content.decode()
        res_dict = json.loads(res)
        token = res_dict.get('access_token')
        refresh_token = res_dict.get('refresh_token')
        return token, refresh_token

    def refresh_token_get(self, refresh_token):
        """
        发送唤醒设备消息
        :param dev_id: 设备云id
        :param msg: 唤醒消息内容
        """
        api_path = '/api/seetong-auth/oauth/token'
        url = f'https://{self.host}{api_path}'
        headers = {
            "User-Agent": "topsee/product-retrospect",
            "Content-Type": "application/json",
            "Authorization": "Basic c2VldG9uZ19jbG91ZF9hZG1pbjpzZWV0b25nX2Nsb3VkX2FkbWluX3NlY3JldA==",
            "Tenant-Id": "000000",
        }

        params = {"grant_type": "refresh_token", "scope": "all", "refresh_token": refresh_token}
        r = requests.post(url, params=params, headers=headers, verify=False)
        res = r.content.decode()
        res_dict = json.loads(res)
        token = res_dict.get('access_token')
        refresh_token = res_dict.get('refresh_token')
        return token, refresh_token

    @refresh_token_wrap()
    def dev_cloud_password_get(self, dev_id):
        api_path = '/api/seetong-device/device-basic-info/get-cloud-password'
        url = f'https://{self.host}{api_path}'
        headers = {
            "User-Agent": "topsee/product-retrospect",
            "Content-Type": "application/json",
            "Authorization": "Basic c2VldG9uZ19jbG91ZF9hZG1pbjpzZWV0b25nX2Nsb3VkX2FkbWluX3NlY3JldA==",
            "Seetong-Auth": self.token,
        }
        params = {"deviceId": dev_id, "password": hashlib.md5(self.password.encode()).hexdigest()}
        r = requests.get(url, params=params, headers=headers, verify=False)
        http_code = r.status_code
        res = r.content.decode()
        res_dict = json.loads(res)
        dev_cloud_password = res_dict.get('data')
        return dev_cloud_password, http_code

    @refresh_token_wrap()
    def device_basic_info(self, dev_sn=None, dev_id=None):
        """查询设备基础信息

        :param dev_sn:设备sn
        :param dev_id: 设备云id
        """
        api_path = '/api/seetong-device/device-basic-info/list'
        url = f'https://{self.host}{api_path}'
        headers = {
            "User-Agent": "topsee/product-retrospect",
            "Content-Type": "application/json",
            "Authorization": "Basic c2VldG9uZ19jbG91ZF9hZG1pbjpzZWV0b25nX2Nsb3VkX2FkbWluX3NlY3JldA==",
            "Seetong-Auth": self.token,
        }
        params = {"deviceSn": dev_sn, "devId": dev_id, "current": "1", "size": "10", "descs": "devId"}
        if not dev_sn:
            del params['deviceSn']
        if not dev_id:
            del params['devId']
        r = requests.get(url, params=params, headers=headers, verify=False)
        http_code = r.status_code
        res = r.content.decode()
        res_dict = json.loads(res)
        return res_dict, http_code

    @refresh_token_wrap()
    def device_basic_info_detail(self, dev_sn):
        """查询设备详细信息
        :param dev_sn: 设备sn
        """

        dev_id = self.dev_id_get(dev_sn=dev_sn)
        api_path = '/api/seetong-device/device-basic-info/detail'
        url = f'https://{self.host}{api_path}'
        headers = {
            "User-Agent": "topsee/product-retrospect",
            "Content-Type": "application/json",
            "Authorization": "Basic c2VldG9uZ19jbG91ZF9hZG1pbjpzZWV0b25nX2Nsb3VkX2FkbWluX3NlY3JldA==",
            "Seetong-Auth": self.token,
        }
        params = {"devId": dev_id}
        r = requests.get(url, params=params, headers=headers, verify=False)
        http_code = r.status_code
        res = r.content.decode()
        res_dict = json.loads(res)
        return res_dict, http_code

    @refresh_token_wrap()
    def device_bind_user(self, dev_sn):
        """查询设备绑定用户信息
        :param dev_sn: 设备sn

        """
        dev_id = self.dev_id_get(dev_sn=dev_sn)
        api_path = '/api/seetong-member-device/device-bind-user/list'
        url = f'https://{self.host}{api_path}'
        headers = {
            "User-Agent": "topsee/product-retrospect",
            "Content-Type": "application/json",
            "Authorization": "Basic c2VldG9uZ19jbG91ZF9hZG1pbjpzZWV0b25nX2Nsb3VkX2FkbWluX3NlY3JldA==",
            "Seetong-Auth": self.token,
        }
        params = {"deviceId": dev_id}
        r = requests.get(url, params=params, headers=headers, verify=False)
        http_code = r.status_code
        res = r.content.decode()
        res_dict = json.loads(res)
        return res_dict, http_code

    def dev_name_get(self, dev_sn):
        """查询设备名称
        :param dev_sn: 设备sn
        """
        #return self.device_bind_user(dev_sn)[0].get('data').get('bindUserList')[0].get('deviceName')
        result = self.device_bind_user(dev_sn)
        bind_user_list = result[0].get('data', {}).get('bindUserList', [])
        if bind_user_list:
            return bind_user_list[0].get('deviceName')
        else:
            # 返回默认值或抛出自定义异常
            return f"未知设备_{dev_sn}"  # 或 raise ValueError(f"设备 {dev_sn} 未绑定用户")

    def dev_version(self, dev_sn):
        """
        _summary_

        :param dev_sn: _description_
        :return: _description_
        """
        return self.device_basic_info_detail(dev_sn)[0].get('data').get('fileVersion')

    def dev_id_get(self, dev_sn):
        """查询设备云id

        :param dev_sn: 设备sn
        """
        return self.device_basic_info(dev_sn=dev_sn)[0].get('data').get('records')[0].get('devId')

    def dev_sn_get(self, dev_id):
        """查询设备sn

        :param dev_id: 设备云id
        """
        return self.device_basic_info(dev_id=dev_id)[0].get('data').get('records')[0].get('devSN')

    @refresh_token_wrap()
    def wake(self, dev_id, msg='test'):
        """
        发送唤醒设备消息
        :param dev_id: 设备云id
        :param msg: 唤醒消息内容
        """
        api_path = '/api/seetong-device-media-command/siot-media/command/send'
        url = f'https://{self.host}{api_path}'
        headers = {
            "User-Agent": "topsee/product-retrospect",
            "Content-Type": "application/json",
            "Authorization": "Basic c2VldG9uZ19jbG91ZF9hZG1pbjpzZWV0b25nX2Nsb3VkX2FkbWluX3NlY3JldA==",
            "Seetong-Auth": self.token,
        }

        data = {'devId': dev_id, 'type': 'LOW_POWER_WAKE', 'msg': msg}
        r = requests.post(url, data=json.dumps(data), headers=headers, verify=False)
        http_code = r.status_code
        res = r.content.decode()
        res_dict = json.loads(res)
        return res, http_code

    def wake_many_times(self, dev_id, msg, count, interval, interval_unit='s'):
        """
        持续发送唤醒消息
        :param dev_id: 设备云id
        :param msg: 唤醒消息内容
        :param count: 循环次数
        :param interval: 循环间隔时间，list类型，随机list中2个数字之间的数字
        """
        t_name = threading.current_thread().name
        for i in range(0, 0 + count):
            msg1 = f'{msg}-{i + 1}'
            res = self.wake(dev_id=dev_id, msg=msg1)
            print(f'{datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")} {t_name} {dev_id}第{i + 1}次唤醒，发送消息：{msg1} 结果: {res}')
            if interval_unit == 's':
                t = random.randint(interval[0], interval[1])
            if interval_unit == 'ms':
                t = random.randint(interval[0], interval[1]) / 1000
            time.sleep(t)

    def wake_test(self, dev_id, msg):
        """_summary_

        :param dev_id: 设备云id
        :param msg: 唤醒消息内容
        :return: 唤醒返回消息
        """
        url = 'http://114.116.200.81:20077/siot-media/command/send'
        headers = {"Content-Type": "application/json"}
        data = {'devId': dev_id, 'type': 'LOW_POWER_WAKE', 'msg': msg}
        r = requests.post(url, data=json.dumps(data), headers=headers)
        res = r.content.decode()
        return res

    def wake_many_times_test(self, dev_id, msg, count, interval):
        """
        持续发送唤醒消息
        :param dev_id: 设备云id
        :param msg: 唤醒消息内容
        :param count: 循环次数
        :param interval: 循环间隔时间，list类型，随机list中2个数字之间的数字
        """
        for i in range(0, 0 + count):
            msg1 = f'{msg}-{i + 1}'
            res = self.wake(dev_id, msg1)
            print(f'{datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")} 第{i + 1}次唤醒，发送消息：{msg1} 结果: {res}')
            t = random.randint(interval[0], interval[1])
            time.sleep(t)

    @refresh_token_wrap()
    def member_list(self, mobile=None, user_name=None, email=None):
        """用户信息
        :param mobile: 手机号
        :param user_name: 用户名
        :param email: 邮箱
        """
        api_path = '/api/seetong-client/client/member/list'
        url = f'https://{self.host}{api_path}'
        headers = {
            "User-Agent": "topsee/product-retrospect",
            "Content-Type": "application/json",
            "Authorization": "Basic c2VldG9uZ19jbG91ZF9hZG1pbjpzZWV0b25nX2Nsb3VkX2FkbWluX3NlY3JldA==",
            "Seetong-Auth": self.token,
        }
        params = {"current": "1", "size": "10", "descs": "id"}
        if mobile:
            params['mobile'] = mobile
        if user_name:
            params['username'] = user_name
        if email:
            params['email'] = email
        r = requests.get(url, params=params, headers=headers, verify=False)
        http_code = r.status_code
        res = r.content.decode()
        res_dict = json.loads(res)
        return res_dict, http_code

    @refresh_token_wrap()
    def user_bind_device_list(self, member_id):
        """
        获取用户绑定设备列表
        :param member_id: 用户ID
        :return: 用户绑定设备列表
        """
        api_path = '/api/seetong-member-device/user-bind-device/list'
        url = f'https://{self.host}{api_path}'
        headers = {
            "User-Agent": "topsee/product-retrospect",
            "Content-Type": "application/json",
            "Authorization": "Basic c2VldG9uZ19jbG91ZF9hZG1pbjpzZWV0b25nX2Nsb3VkX2FkbWluX3NlY3JldA==",
            "Seetong-Auth": self.token,
        }
        params = {"userId": member_id}
        r = requests.get(url, params=params, headers=headers, verify=False)
        http_code = r.status_code
        res = r.content.decode()
        res_dict = json.loads(res)
        return res_dict, http_code

    @refresh_token_wrap()
    def tencent_cloud_storage_device_query(self, dev_sn):
        api_path = '/api/seetong-traceable/device-relation-tencent/detail/by/device-sn'
        url = f'https://{self.host}{api_path}'
        headers = {
            "User-Agent": "topsee/product-retrospect",
            "Content-Type": "application/json",
            "Authorization": "Basic c2VldG9uZ19jbG91ZF9hZG1pbjpzZWV0b25nX2Nsb3VkX2FkbWluX3NlY3JldA==",
            "Seetong-Auth": self.token,
        }
        params = {"deviceSn": dev_sn}
        r = requests.get(url, params=params, headers=headers, verify=False)
        http_code = r.status_code
        res = r.content.decode()
        res_dict = json.loads(res)
        return res_dict, http_code

    @refresh_token_wrap()
    def tencent_cloud_storage_time_aixs(self, dev_sn, date):
        api_path = '/api/seetong-mall/va-service-manage/tencent-cloud-storage/describeCloudStorageTime'
        url = f'https://{self.host}{api_path}'
        headers = {
            "User-Agent": "topsee/product-retrospect",
            "Content-Type": "application/json",
            "Authorization": "Basic c2VldG9uZ19jbG91ZF9hZG1pbjpzZWV0b25nX2Nsb3VkX2FkbWluX3NlY3JldA==",
            "Seetong-Auth": self.token,
        }
        tencent_cloud_storage_device, _ = self.tencent_cloud_storage_device_query(dev_sn)
        start_time = int(time.mktime(time.strptime(date, "%Y-%m-%d")))
        end_time = start_time + 24 * 60 * 60 - 1
        device_name, product_id = tencent_cloud_storage_device.get('data').get('tencentDeviceName'), tencent_cloud_storage_device.get('data').get('tencentProductId')
        params = {"device_name": device_name, "product_id": product_id, "date": date, "startTime": start_time, "endTime": end_time}
        r = requests.get(url, params=params, headers=headers, verify=False)
        http_code = r.status_code
        res = r.content.decode()
        res_dict = json.loads(res)
        return res_dict, http_code

    def query_device(self, mobile=None, user_name=None, email=None, model=None, account_flag=False, sn_flag=False):
        """打印用户绑定设备信息
        :param mobile: 手机号
        :param user_name: 用户名
        :param email: 邮箱
        :param write_flag: 是否保存到csv文件
        """
        res, _ = self.member_list(mobile=mobile, user_name=user_name, email=email)
        member_id = res.get('data').get('records')[0].get('id')
        res, _ = self.user_bind_device_list(member_id)
        dev_info = res.get('data')
        dev_account_list = []
        dev_sn_list = []
        for dev in dev_info:
            dev_name = dev.get('deviceName')
            dev_sn = dev.get('deviceSn')
            dev_id = dev.get('deviceId')
            dev_cloud_password, _ = self.dev_cloud_password_get(dev_id)
            dev_version = self.dev_version(dev_sn)
            dev_account = f'{dev_name},{dev_sn},{dev_id},{dev_cloud_password},{dev_version}'
            # dev_account = [dev_name, dev_sn, dev_id, dev_cloud_password, dev_version]
            if (not model) or (model and dev_version.find(model) != -1):
                dev_account_list.append(dev_account)
                dev_sn_list.append(dev_sn)
                # print(f'{dev_name} {dev_sn} {dev_id} {dev_cloud_password} {dev_version}')
                print(dev_account)

        if account_flag:
            current_path = os.path.dirname(os.path.abspath(__file__))
            with open(current_path + os.sep + 'syj.csv', 'w', encoding='utf-8') as f:
                for dev in dev_account_list:
                    f.writelines(dev + '\n')
        if sn_flag:
            config_json = self.update_config_creat(dev_sn_list)
            print(config_json)

    def update_config_creat(self, dev_sn_list):
        config_dict = {}
        records = []
        for dev_sn in dev_sn_list:
            dev_id = self.dev_id_get(dev_sn)
            dev_login_user = "admin"
            dev_login_pass, _ = self.dev_cloud_password_get(dev_id)
            dev_channel_count = "2"
            dev_channel_state = "11"

            record = {"DevSN": dev_sn, "DevID": dev_id, "DevLoginUser": dev_login_user, "DevLoginPass": dev_login_pass, "DevChannelCount": dev_channel_count, "DevChannelState": dev_channel_state}
            records.append(record)
        config_dict['RECORDS'] = records
        config_json = json.dumps(config_dict, indent=4)
        return config_json

    def cloud_id_to_ipc(self, cloud_id_list, write_flag=False):
            """
            传入云ID的列表，输出设备的详细信息
            :param cloud_id_list: IPC的云id列表
            :param write_flag: 是否保存到csv文件,True保存，False不保存，默认不保存
            """
            dev_account_list = []
            for i in cloud_id_list:
                sn = self.dev_sn_get(i)
                name = self.dev_name_get(sn)
                dev_cloud_password, _ = self.dev_cloud_password_get(i)
                dev_account = f'{name},{sn},{i},{dev_cloud_password}'
                print(dev_account)
                dev_account_list.append(dev_account)

            if write_flag is True:
                current_path = os.path.dirname(os.path.abspath(__file__))
                with open(current_path + os.sep + 'syj.csv', 'w', encoding='utf-8') as f:
                    for dev in dev_account_list:
                        f.writelines(dev + '\n')

    def cloud_sn_to_ipc(self, cloud_sn_list, write_flag=True):
            """
            传入sn的列表，输出设备的详细信息
            :param cloud_sn_list: IPC的sn列表
            :param write_flag: 是否保存到csv文件,True保存，False不保存，默认不保存
            """
            dev_account_list = []
            for i in cloud_sn_list:
                cloud_id = self.dev_id_get(i)
                name = self.dev_name_get(i)
                dev_cloud_password, _ = self.dev_cloud_password_get(cloud_id)
                dev_account = f'{name},{i},{cloud_id},{dev_cloud_password}'
                print(dev_account)
                dev_account_list.append(dev_account)

            if write_flag is True:
                current_path = os.path.dirname(os.path.abspath(__file__))
                with open(current_path + os.sep + 'syj.csv', 'w', encoding='utf-8') as f:
                    for dev in dev_account_list:
                        f.writelines(dev + '\n')

    def read_sn_from_xlsx(self, file_path):
        """
        从xlsx文件中自动查找包含'sn'的列（不区分大小写），读取该列数据
        :param file_path: xlsx文件路径
        :return: SN列表
        """
        wb = load_workbook(file_path)
        sheet = wb.active

        # 获取第一行，查找SN列
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        sn_col_index = None
        for idx, cell in enumerate(header_row):
            if cell and 'sn' in str(cell).lower():
                sn_col_index = idx
                break

        if sn_col_index is None:
            raise ValueError("未找到包含 'sn' 的列")

        sn_list = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始
            if len(row) > sn_col_index and row[sn_col_index]:
                sn_list.append(str(row[sn_col_index]).strip())
        return sn_list

if __name__ == '__main__':
    username = 'yinjia'
    password = 'Yjtest123456.'
    dev_id = 36240360
    msg = 'wake test'
    env = 'pro'
    time_now = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    app = Console_Admin_System(env, username, password)
    # app.wake(dev_id)
    # app.wake_many_times(dev_id, msg, 10, [1, 1])
    # app.data("0C5C94C4A8923C36")
    # app.wake_many_times(dev_id, msg, 10, [1, 1])
    # dev_id = app.dev_id_get(dev_sn)

    # current_path = os.path.dirname(os.path.abspath(__file__))
    # dev_file = current_path + os.sep + 'data.txt'
    # dev_info_list = dev_info_get(dev_file)

    # app.query_device(mobile='13553728129', model='TD53E10', account_flag=True, sn_flag=True)
    # app.query_device(mobile='18025358468', model='TD53E10', account_flag=True, sn_flag=True)
    # app.query_device(user_name='ruibin.liu@tpsee.com', model='TD53E10', account_flag=True, sn_flag=True)
    # app.query_device(user_name='ipctest03', account_flag=True, sn_flag=True)
    #app.query_device(user_name='18000233534', model='TD53E30', account_flag=True, sn_flag=False)
    # app.query_device(user_name='kjtest02', model='TD53E10', account_flag=True, sn_flag=False)

    # res, _ = app.tencent_cloud_storage_time_aixs('0C2D4257A4963A46', '2025-03-31')
    # time_list = res.get('data').get('timeList')
    # total = 0
    # for time_aixs in time_list:
    #     total += time_aixs.get('endtime') - time_aixs.get('starttime') + 1
    # lost_time = 24 * 60 * 60 - total
    # print(lost_time)

    # 从xlsx文件中读取SN列表
    sn_list = app.read_sn_from_xlsx('E10-20260106-1052pcs.xlsx')
    print(f"从xlsx读取到 {len(sn_list)} 个SN")
    app.cloud_sn_to_ipc(cloud_sn_list=sn_list)

    # 注释掉硬编码的调用
    # app.cloud_sn_to_ipc(cloud_sn_list=["0CA6508CAED13D4A","0C97099AAE5E3D8F","0C99BF69AE883D40","0C91512EAE513D58","0C2792A6EC8C3A1D","0CA2248AAEC43D0B","0C9EE8B7AE9A3DCE","0CAB0D68AEDE3DF5","0CAB2F3EAEDF3DEE","0CA6554BAED13D0E","0CA66443AED23D16","0C99A239AE873DF2","0C9D537FAE983DFE","0C9598E2AE5B3D61","0C91874DAE533DAF","0C9ADC5CAE8D3D56","0CA1F7EFAEC33D41","0C93E63FAE553D04","0C9C065CAE903D85","0CA6670BAED23DE1","0C9AE798AE8D3D9D","0C998CF3AE873D96","0C995EE9AE853D5C","0C915AC7AE513DFA","0C963B8FAE5B3DB2","0C91AFD2AE543D5D","0C919D0CAE543D85","0C919D23AE543D9C","0C93E3A9AE553D6B","0C9C065CAE903D85","0C96472BAE5C3D5B","0C99A819AE873DD8","0C97A003AE833DB4","0C91A666AE543DE8","0C9ABD4DAE8D3D28","0C99832BAE863DC4"])
